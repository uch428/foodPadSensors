#!/usr/bin/env python
import rospy
from rosserial_arduino.msg import Adc
import openpyxl as px

wb = px.Workbook()
ws = wb.active
measure = px.Workbook()

r = 2
c = 2

def callback(data):
	rospy.loginfo('adc0: %s', data.adc0)
	rospy.loginfo('adc1: %s', data.adc1)
	rospy.loginfo('adc2: %s', data.adc2)
	rospy.loginfo('adc3: %s', data.adc3)
	rospy.loginfo('adc4: %s', data.adc4)
	rospy.loginfo('adc5: %s', data.adc5)
	rospy.loginfo('adc6: %s', data.adc6)
	rospy.loginfo('adc7: %s', data.adc7)
	#print(data.adc0)

	print('---')
	
	csvWrite(data)


def csvWrite(data):
	global r
	global c
	

	global ws
	ws.cell(row = r, column = c, value = data.adc0)
	ws.cell(row = r, column = c+1, value = data.adc1)
	ws.cell(row = r, column = c+2, value = data.adc2)
	ws.cell(row = r, column = c+3, value = data.adc3)
	ws.cell(row = r, column = c+4, value = data.adc4)
	ws.cell(row = r, column = c+5, value = data.adc5)
	ws.cell(row = r, column = c+6, value = data.adc6)
	ws.cell(row = r, column = c+7, value = data.adc7)


	r += 1
	c = 2
	#c = 2

def listener():
	rospy.init_node('listener', anonymous = True)
	print('========================')
	rospy.Subscriber('adc', Adc, callback)
	rospy.spin()
	print('=====')

if __name__ == '__main__':
	

	label = ['adc0', 'adc1', 'adc2', 'adc3', 'adc4', 'adc5', 'adc6', 'adc7'] 
	i = 1
	for w in label:
		ws.cell(row = 1, column = i+1, value = w)	
		i += 1
	listener()

	wb.save('/home/yuta/ros/workspaces/myWorkspace/data1.xlsx')
	#bk.save('data11.xlsx')



# answers.ros.org/question/265150/need-help-in-subscribing-from-one-topic-and-publishing-to-another/
